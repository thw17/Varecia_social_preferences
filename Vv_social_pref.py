# Tim Webster, University of Utah, 2019
# This script runs a handful of analyses for our paper, Baden et al. (2019 and revision in 2020)
# It is meant to be a one-time script and thus, the code is functional but isn't
# very clean or concise. It aims to run randomizations analyses for
# association and relatedness (are top associates more related than expected?) and
# for UDOI and relatedness (are dyads with highest UDOI values more related than expected?).
# In addition, it creates one of the plots in the paper.

## Final command line used: python ~/Projects/Github_repos/Varecia_social_preferences/Vv_social_pref.py --association vv_association_matrix.xlsx --pref_matrix vv_sig_pref_assoc_matrix.xlsx --categories vv_category_matrix.xlsx --udoi vv_rangeoverlap_matrix.xlsx --permutations 10000 --relatedness vv_relatedness_matrix_no_missing_data.xlsx

import argparse
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import seaborn as sns
from sklearn.linear_model import LogisticRegression
from scipy import stats
from openpyxl import load_workbook


def parse_args():
	""" Parse command line arguments """
	parser = argparse.ArgumentParser(
		description="")

	parser.add_argument(
		"--association", required=True,
		help="Input association matrix Excel file")

	parser.add_argument(
		"--pref_matrix", required=True,
		help="Input preferred associate matrix Excel file (where preferences were determined using another program)")

	parser.add_argument(
		"--categories", required=True,
		help="Input category matrix Excel file")

	parser.add_argument(
		"--relatedness", required=True,
		help="Input relatedness matrix Excel file")

	parser.add_argument(
		"--udoi", required=True,
		help="Input home range overlap (UDOI) matrix Excel file")

	parser.add_argument(
		"--permutations", default=1000, type=int,
		help="Number of permutations.")

	args = parser.parse_args()
	return args

def read_excel_matrix_as_edgelist(file1):
	wb = load_workbook(filename = file1)
	sheet = wb.active
	d1 = {}
	id_idx = []
	row_num = 0
	row_id = None
	for row in sheet.iter_rows():
		cell_num = 0
		for cell in row:
			if row_num == 0 and cell_num == 0:
				cell_num += 1
				continue
			else:
				if row_num == 0:
					id_idx.append(cell.value)
				else:
					if cell_num == 0:
						row_id = cell.value
					else:
						col_id = id_idx[cell_num - 1]
						if col_id == row_id:
							pass
						else:
							if (col_id, row_id) not in d1:
								d1[(col_id, row_id)] = cell.value
				cell_num += 1
		row_num += 1
	return d1

def main():
	args = parse_args()

	assoc = read_excel_matrix_as_edgelist(args.association)
	pref_assoc = read_excel_matrix_as_edgelist(args.pref_matrix)
	relate = read_excel_matrix_as_edgelist(args.relatedness)
	udoi = read_excel_matrix_as_edgelist(args.udoi)

	master_keys = []
	for i in assoc:
		master_keys.append(i)
	for i in relate:
		master_keys.append(i)
	for i in udoi:
		master_keys.append(i)

	master_keys_unique = []
	for i in master_keys:
		if i not in master_keys_unique:
			if (i[1], i[0]) not in master_keys_unique:
				master_keys_unique.append(i)

	assoc_temp = {}
	for k in assoc:
		if (k[1], k[0]) in master_keys_unique:
			assoc_temp[k] = assoc[k]

	pref_assoc_temp = {}
	for k in pref_assoc:
		if (k[1], k[0]) in master_keys_unique:
			pref_assoc_temp[k] = pref_assoc[k]

	udoi_temp = {}
	for k in udoi:
		if (k[1], k[0]) in master_keys_unique:
			udoi_temp[k] = udoi[k]

	relate_temp = {}
	for k in relate:
		if (k[1], k[0]) in master_keys_unique:
			relate_temp[k] = relate[k]

	print("Association dyads: {}".format(len(assoc_temp)))
	print("Preference dyads: {}".format(len(pref_assoc_temp)))
	print("UDOI dyads: {}".format(len(udoi_temp)))
	print("Relatedness dyads: {}".format(len(relate_temp)))
	print("")

	print("######################################################################")
	print("Running analyses for original paper submission, with preference based on percentile")
	print("--these analyses were based on Carter et al. 2013 and Best et al. 2014")
	print("--they involved assessing preferences based on percentile of indices")
	print("--during paper revisions, at the suggest on of a reviewer, we abandoned this approach for linear models (see below)")
	print("")

	# Set up dictionaries for dyads in both association and relatedness matrices
	assoc_ak = {k: assoc_temp[k] for k in relate if k in assoc_temp}
	relate_ak = {k: relate[k] for k in assoc_ak if k in relate}

	assoc_ak_values = list(assoc_ak.values())
	relate_ak_values = list(relate_ak.values())

	# Set up dictionaries for dyads in both udoi and relatedness matrices
	udoi_uk = {k: udoi_temp[k] for k in relate if k in udoi_temp}
	relate_uk = {k: relate[k] for k in udoi_uk if k in relate}

	udoi_uk_values = list(udoi_uk.values())
	relate_uk_values = list(relate_uk.values())

	# Look for preferred, casual, and avoided dyads
	# Note dyads used for association and UDOI, respectively, only include those
	# for which there is relatedness data

	a95 = np.percentile(np.array(assoc_ak_values), 95)
	u95 = np.percentile(np.array(udoi_uk_values), 95)

	a5 = np.percentile(np.array(assoc_ak_values), 5)
	u5 = np.percentile(np.array(udoi_uk_values), 5)

	a_pref_count = 0
	a_casual_count = 0
	a_avoid_count = 0
	for x in assoc_ak_values:
		if x > a95:
			a_pref_count += 1
		elif x < a95:
			a_avoid_count += 1
		else:
			a_casual_count += 1
	a_other_count = a_casual_count + a_avoid_count

	u_pref_count = 0
	u_casual_count = 0
	u_avoid_count = 0
	for x in udoi_uk_values:
		if x > u95:
			u_pref_count += 1
		elif x < u5:
			u_avoid_count += 1
		else:
			u_casual_count += 1
	u_other_count = u_casual_count + u_avoid_count

	# Print preferred, casual, avoid counts. Association matrix very sparse, so
	# just focus on preferred vs. not afterward

	print("Association counts (number of dyads) for preferred, casual, avoided:")
	print("{}, {}, {}".format(a_pref_count, a_casual_count, a_avoid_count))
	print("")

	print("UDOI counts (number of dyads) for preferred, casual, avoided:")
	print("{}, {}, {}".format(u_pref_count, u_casual_count, u_avoid_count))
	print("")

	# Print preferred associates (note, run on arrays and not pandas DF because
	# only need to make sure dyad present in assoc and relate, not udoi)
	temp_pref = []
	temp_not = []

	print("Preferred associates (no particular order):")
	for pair in assoc_ak:
		if assoc_ak[pair] > a95:
			temp_pref.append(relate_ak[pair])
			print("{}: {}".format(pair, assoc_ak[pair]))
		else:
			temp_not.append(relate_ak[pair])

	empirical_pref_a = np.mean(temp_pref)
	empirical_not_a = np.mean(temp_not)

	print("Preferred mean: {}".format(empirical_pref_a))
	print("Not preferred mean: {}".format(empirical_not_a))
	print("")

	# Print preferred udoi (note, run on arrays and not pandas DF because
	# only need to make sure dyad present in udoi and relate, not assoc)
	temp_pref = []
	temp_not = []

	print("'Preferred' UDOI (no particular order):")
	for pair in udoi_uk:
		if udoi_uk[pair] > u95:
			temp_pref.append(relate_uk[pair])
			print("{}: {}".format(pair, udoi_uk[pair]))
		else:
			temp_not.append(relate_uk[pair])

	empirical_pref_u = np.mean(temp_pref)
	empirical_not_u = np.mean(temp_not)

	print("Preferred mean: {}".format(empirical_pref_u))
	print("Not preferred mean: {}".format(empirical_not_u))
	print("")

	### Association and relatedness randomizations
	#without replacement
	pref_dist = []
	not_dist = []
	for i in range(0, args.permutations):
		k = np.random.permutation(relate_ak_values)
		pref1 = k[:a_pref_count]
		not1 = k[a_pref_count:]
		pref_dist.append(np.mean(pref1))
		not_dist.append(np.mean(not1))

	#with replacement
	pref_with_repl = []
	not_with_repl = []
	for i in range(0, args.permutations):
		pref_with_repl.append(np.mean(np.random.choice(relate_ak_values, a_pref_count)))
		not_with_repl.append(np.mean(np.random.choice(relate_ak_values, a_other_count)))

	# Print
	# print(
	# 	"Association--Emprical preferred mean: {}. Percentile of {} permutations without replacement: {}".format(
	# 		empirical_pref_a, args.permutations, stats.percentileofscore(
	# 			pref_dist, empirical_pref_a, kind="strict")))
	# print(
	# 	"Association--Emprical not preferred mean: {}. Percentile of {} permutations without replacement: {}".format(
	# 		empirical_not_a, args.permutations, stats.percentileofscore(
	# 			not_dist, empirical_not_a, kind="strict")))
	print(
		"Association--Emprical preferred mean: {}. Percentile of {} permutations with replacement: {}".format(
			empirical_pref_a, args.permutations, stats.percentileofscore(
				pref_with_repl, empirical_pref_a, kind="strict")))
	print(
		"Association--Emprical not preferred mean: {}. Percentile of {} permutations with replacement: {}".format(
			empirical_not_a, args.permutations, stats.percentileofscore(
				not_with_repl, empirical_not_a, kind="strict")))
	print("")


	### UDOI and relatedness randomizations
	#without replacement
	pref_dist = []
	not_dist = []
	for i in range(0, args.permutations):
		k = np.random.permutation(relate_uk_values)
		pref1 = k[:u_pref_count]
		not1 = k[u_pref_count:]
		pref_dist.append(np.mean(pref1))
		not_dist.append(np.mean(not1))

	#with replacement
	pref_with_repl = []
	not_with_repl = []
	for i in range(0, args.permutations):
		pref_with_repl.append(np.mean(np.random.choice(relate_uk_values, u_pref_count)))
		not_with_repl.append(np.mean(np.random.choice(relate_uk_values, u_other_count)))

	# Print
	# print(
	# 	"UDOI--Emprical preferred mean: {}. Percentile of {} permutations without replacement: {}".format(
	# 		empirical_pref_u, args.permutations, stats.percentileofscore(
	# 			pref_dist, empirical_pref_u, kind="strict")))
	# print(
	# 	"UDOI--Emprical not preferred mean: {}. Percentile of {} permutations without replacement: {}".format(
	# 		empirical_not_u, args.permutations, stats.percentileofscore(
	# 			not_dist, empirical_not_u, kind="strict")))
	print(
		"UDOI--Emprical preferred mean: {}. Percentile of {} permutations with replacement: {}".format(
			empirical_pref_u, args.permutations, stats.percentileofscore(
				pref_with_repl, empirical_pref_u, kind="strict")))
	print(
		"UDOI--Emprical not preferred mean: {}. Percentile of {} permutations with replacement: {}".format(
			empirical_not_u, args.permutations, stats.percentileofscore(
				not_with_repl, empirical_not_u, kind="strict")))
	print("")


	## Rerun analyses for preferred associates where preferences where determined
	## Using an external program (SOCPROG)
	## This code uses routines run above and thus, the following code cannot be
	## extracted and run on its own

	print("######################################################################")
	print("Now conducting analyses using preferred associates defined externally")
	print("--these are the updated analyses used instead of the percentile method described above")
	print("--we first repeat the permutation test with the new preferences")
	print("--we next use logistic regression with randomizations, which is what we ended up using in the paper")
	print("")

	pref_assoc_ak = {k: pref_assoc_temp[k] for k in assoc_ak if k in pref_assoc_temp}

	# Ensure values for association and relatedness are in same order
	pref_assoc_ak_values = []
	relate_ak_values = []
	for k in pref_assoc_ak:
		pref_assoc_ak_values.append(pref_assoc_ak[k])
		relate_ak_values.append(relate_ak[k])

	# pref_assoc_ak_values = list(pref_assoc_ak.values())
	# relate_ak_values = list(relate_ak.values())

	a_pref_count = 0
	a_other_count = 0
	for x in pref_assoc_ak_values:
		if x > a95:
			a_pref_count += 1
		else:
			a_other_count += 1

	temp_pref = []
	temp_not = []

	print("Preferred associates (no particular order):")
	for pair in pref_assoc_ak:
		if pref_assoc_ak[pair] == 1:
			temp_pref.append(relate_ak[pair])
			print("{}: {}".format(pair, pref_assoc_ak[pair]))
		else:
			temp_not.append(relate_ak[pair])

	empirical_pref_a = np.mean(temp_pref)
	empirical_not_a = np.mean(temp_not)

	print("Preferred mean: {}".format(empirical_pref_a))
	print("Not preferred mean: {}".format(empirical_not_a))
	print("")

	pref_dist = []
	not_dist = []
	for i in range(0, args.permutations):
		k = np.random.permutation(relate_ak_values)
		pref1 = k[:a_pref_count]
		not1 = k[a_pref_count:]
		pref_dist.append(np.mean(pref1))
		not_dist.append(np.mean(not1))

	pref_with_repl = []
	not_with_repl = []
	for i in range(0, args.permutations):
		pref_with_repl.append(np.mean(np.random.choice(relate_ak_values, a_pref_count)))
		not_with_repl.append(np.mean(np.random.choice(relate_ak_values, a_other_count)))

	print(
		"Association--Emprical preferred mean: {}. Percentile of {} permutations with replacement: {}".format(
			empirical_pref_a, args.permutations, stats.percentileofscore(
				pref_with_repl, empirical_pref_a, kind="strict")))
	print(
		"Association--Emprical not preferred mean: {}. Percentile of {} permutations with replacement: {}".format(
			empirical_not_a, args.permutations, stats.percentileofscore(
				not_with_repl, empirical_not_a, kind="strict")))
	print("")


	print("")
	print("Logistic regression")
	model = LogisticRegression().fit(np.array(relate_ak_values).reshape(-1,1), pref_assoc_ak_values)
	c1 = model.coef_
	i1 = model.intercept_
	print("Model coef: {}".format(c1))
	print("Model intercept: {}".format(i1))

	coef_dist = []
	for i in range(0, args.permutations):
		k = np.random.permutation(relate_ak_values)
		temp_model = LogisticRegression().fit(np.array(k).reshape(-1,1), pref_assoc_ak_values)
		coef_dist.append(temp_model.coef_)

	print(
		"Logistic regression permutation analysis p-value: {}".format(
			1.0 - stats.percentileofscore(
				coef_dist, c1, kind="strict")/100))



	#########################################################################
	# Scatterplot

	# Get categories
	categ = read_excel_matrix_as_edgelist(args.categories)

	# Create combined dictionary and then convert to pandas dataframe

	assoc1 = {k: assoc_ak[k] for k in udoi_uk if k in assoc_ak}
	relate1 = {k: relate_ak[k] for k in assoc1 if k in relate}
	udoi1 = {k: udoi_uk[k] for k in assoc1 if k in udoi}
	categ1 = {k: categ[k] for k in assoc1 if k in categ}

	combined_dict = {"dyad": [], "ai": [], "relate":[], "udoi": [], "Category": []}
	for k in assoc1:
		combined_dict["dyad"].append(k)
		combined_dict["ai"].append(assoc1[k])
		combined_dict["relate"].append(relate1[k])
		combined_dict["udoi"].append(udoi1[k])
		combined_dict["Category"].append(categ1[k])

	df = pd.DataFrame.from_dict(combined_dict)
	df = df[["dyad", "ai", "udoi", "relate", "Category"]]
	df["Category2"] = np.where(df["relate"] > 0.25, "Kin", "Nonkin")

	# # plot
	# g = sns.scatterplot(
	# 	x="udoi", y="ai", hue="Category2", palette=["black", "white"],
	# 	data=df, style="Category2", markers=["^", "o"], edgecolor=["white", "black"])
	# handles, labels = g.get_legend_handles_labels()
	# g.legend(handles=handles[1:], labels=labels[1:])
	# g.set(xlabel="Home Range Overlap (UDOI)", ylabel="Association Indices")
	# g2 = g.get_figure()
	# g2.savefig("vv_scatter.pdf", dpi=500, transparent=True)

	# try subplots 9 x 6 inches
	df_kin = df[df["Category2"] == "Kin"]
	df_nonkin = df[df["Category2"] == "Nonkin"]
	g = plt.figure(figsize=[9,6])
	ax1 = g.add_subplot(111)
	sns.scatterplot(
		x="udoi", y="ai", hue="Category2", palette=["black"], data=df_nonkin,
		style="Category2", markers=["^"], edgecolor=["white"], ax=ax1)

	ax2 = g.add_subplot(111, frameon=False, sharex=ax1, sharey=ax1)
	sns.scatterplot(
		x="udoi", y="ai", hue="Category2", palette=["white"], data=df_kin,
		style="Category2", markers=["o"], edgecolor=["black"], ax=ax2)

	# handles, labels = ax1.get_legend_handles_labels()
	# ax1.legend(handles=handles[1:], labels=labels[1:])
	# handles, labels = ax2.get_legend_handles_labels()
	# ax2.legend(handles=handles[1:], labels=labels[1:])
	ax1.set(xlabel="Home Range Overlap (UDOI)", ylabel="Association Indices")
	ax2.set_xlabel('')
	ax2.set_ylabel('')
	ax1.get_legend().remove()
	ax2.get_legend().remove()
	# g2 = g.get_figure()
	g.savefig("vv_scatter_9x6.pdf", dpi=500, transparent=True)

	# try subplots 6 x 4 inches
	df_kin = df[df["Category2"] == "Kin"]
	df_nonkin = df[df["Category2"] == "Nonkin"]
	g = plt.figure(figsize=[6,4])
	ax1 = g.add_subplot(111)
	sns.scatterplot(
		x="udoi", y="ai", hue="Category2", palette=["black"], data=df_nonkin,
		style="Category2", markers=["^"], edgecolor=["white"], ax=ax1)

	ax2 = g.add_subplot(111, frameon=False, sharex=ax1, sharey=ax1)
	sns.scatterplot(
		x="udoi", y="ai", hue="Category2", palette=["white"], data=df_kin,
		style="Category2", markers=["o"], edgecolor=["black"], ax=ax2)

	# handles, labels = ax1.get_legend_handles_labels()
	# ax1.legend(handles=handles[1:], labels=labels[1:])
	# handles, labels = ax2.get_legend_handles_labels()
	# ax2.legend(handles=handles[1:], labels=labels[1:])
	ax1.set(xlabel="Home Range Overlap (UDOI)", ylabel="Association Indices")
	ax2.set_xlabel('')
	ax2.set_ylabel('')
	ax1.get_legend().remove()
	ax2.get_legend().remove()
	# g2 = g.get_figure()
	g.savefig("vv_scatter_6x4.pdf", dpi=500, transparent=True)

if __name__ == "__main__":
	main()
